import openpyxl

# Load workbook
wb = openpyxl.load_workbook('public/INCENTIVE DL SI V1 - Modify 11.12 (1).xlsx', data_only=False)

# Focus on Data sheet
ws = wb['Data']

print("=== PHÂN TÍCH CHI TIẾT SHEET DATA ===\n")

# Bảng 1: Tỷ lệ thưởng theo Kỳ hạn (Rows 3-10, Cols G-K)
print("=== BẢNG 1: TỶ LỆ THƯỞNG THEO KỲ HẠN ===")
for row in range(2, 12):
    cells = []
    for col in range(6, 12):  # F to K
        cell = ws.cell(row=row, column=col)
        val = cell.value if cell.value else ""
        cells.append(str(val)[:15])
    print(f"R{row:02d}: {' | '.join(cells)}")

print("\n=== BẢNG 2: CHỈ TIÊU VÀ TIỀN THƯỞNG ===")
# Rows 14-18, Cols G-L
for row in range(14, 19):
    cells = []
    for col in range(7, 14):  # G to M
        cell = ws.cell(row=row, column=col)
        val = cell.value if cell.value else ""
        cells.append(str(val)[:12])
    print(f"R{row:02d}: {' | '.join(cells)}")

print("\n=== BẢNG 3: BẢO HIỂM THEO DSGN VÀ KỲ HẠN ===")
# Rows 22-28, Cols G-L
for row in range(22, 32):
    cells = []
    for col in range(7, 14):  # G to M
        cell = ws.cell(row=row, column=col)
        val = cell.value if cell.value else ""
        cells.append(str(val)[:15])
    print(f"R{row:02d}: {' | '.join(cells)}")

print("\n=== BẢNG 4: HỆ SỐ PR3/PR6 ===")
# Rows 32-45, Cols G-K
for row in range(32, 52):
    cells = []
    for col in range(7, 14):  # G to M
        cell = ws.cell(row=row, column=col)
        val = cell.value if cell.value else ""
        cells.append(str(val)[:15])
    print(f"R{row:02d}: {' | '.join(cells)}")

# Also check INCENTIVE sheet for main formulas
print("\n\n=== PHÂN TÍCH INCENTIVE SHEET ===")
ws2 = wb['INCENTIVE']
for row in range(1, 40):
    cells = []
    for col in range(1, 10):
        cell = ws2.cell(row=row, column=col)
        val = cell.value
        if val is not None:
            if isinstance(val, str) and val.startswith('='):
                cells.append(f"[{val[:80]}]")
            else:
                cells.append(str(val)[:25])
        else:
            cells.append("")
    if any(cells):
        print(f"R{row:02d}: {' | '.join(cells)}")
